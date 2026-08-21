#!/usr/bin/env python3
"""Explicitly synchronize the provided operation indicator MOCK workbook."""

from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from query import load_config


def literal(value):
    if value is None:
        return "NULL"
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    encoded = str(value).encode("utf-8").hex()
    return f"CONVERT(0x{encoded} USING utf8mb4)"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")
    config = load_config()
    user = os.environ.get("BRRP_DB_ADMIN_USER")
    password = os.environ.get("BRRP_DB_ADMIN_PASSWORD")
    if not user or not password:
        raise SystemExit("同步需要临时设置 BRRP_DB_ADMIN_USER 和 BRRP_DB_ADMIN_PASSWORD")
    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = workbook["指标数据"]
    rows = [tuple(values[:17]) for values in sheet.iter_rows(min_row=5, values_only=True) if values[0]]
    insert_prefix = """
        INSERT INTO operation_indicator_fact
        (period,object_id,object_code,object_name,display_order,indicator_code,indicator_name,
         indicator_category,indicator_type,data_source,data_type,row_position,unit,current_value,
         previous_value,yoy_amount,yoy_rate)
        VALUES
    """
    update_suffix = """
        ON DUPLICATE KEY UPDATE
          object_code=VALUES(object_code),object_name=VALUES(object_name),display_order=VALUES(display_order),
          indicator_name=VALUES(indicator_name),indicator_category=VALUES(indicator_category),
          indicator_type=VALUES(indicator_type),data_source=VALUES(data_source),data_type=VALUES(data_type),
          row_position=VALUES(row_position),unit=VALUES(unit),current_value=VALUES(current_value),
          previous_value=VALUES(previous_value),yoy_amount=VALUES(yoy_amount),yoy_rate=VALUES(yoy_rate),deleted_flag=0
    """
    statements = ["START TRANSACTION;"]
    for offset in range(0, len(rows), 200):
        values = ",\n".join("(" + ",".join(literal(value) for value in row) + ")" for row in rows[offset:offset + 200])
        statements.append(insert_prefix + values + update_suffix + ";")
    statements.extend(["COMMIT;", "SELECT COUNT(*) AS active_fact_rows FROM operation_indicator_fact WHERE deleted_flag=0;"])
    mysql = shutil.which("mysql")
    if not mysql:
        raise SystemExit("未找到 mysql 命令行客户端")
    env = os.environ.copy(); env["MYSQL_PWD"] = password
    result = subprocess.run(
        [mysql, "--host", config["db_host"], "--port", str(config["db_port"]), "--user", user,
         "--database", config["db_name"], "--default-character-set=utf8mb4", "--batch", "--raw"],
        input="\n".join(statements), capture_output=True, text=True, encoding="utf-8", env=env, check=False,
    )
    if result.returncode != 0:
        raise SystemExit(result.stderr.strip() or "同步失败")
    count = result.stdout.splitlines()[-1] if result.stdout.splitlines() else "unknown"
    print(f"Synchronized {len(rows)} workbook rows; active fact rows: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
